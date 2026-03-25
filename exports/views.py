from io import BytesIO
from urllib.parse import quote

from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
from django.views.generic import TemplateView, View

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from inventory.models import Building, GlobalStock, BuildingStock

from .forms import ExportIssuesForm, ExportStocksForm
from .services import build_transactions_workbook


class ExportHomeView(LoginRequiredMixin, TemplateView):
    template_name = "exports/export_form.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["issues_form"] = ExportIssuesForm()
        ctx["stocks_form"] = ExportStocksForm()
        return ctx


def _make_excel_response(workbook, filename: str) -> HttpResponse:
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    encoded_filename = quote(filename)
    response["Content-Disposition"] = f"attachment; filename*=UTF-8''{encoded_filename}"
    return response


def _balance_suffix(balance_type: str) -> str:
    mapping = {
        "all": "all",
        "balance": "balance",
        "non_balance": "non_balance",
        "balance_only": "balance",
        "non_balance_only": "non_balance",
    }
    return mapping.get(balance_type, "all")


class ExportTransactionsExcelView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        form = ExportIssuesForm(request.POST)
        if not form.is_valid():
            return HttpResponse("Некорректные параметры выгрузки.", status=400)

        date_from = form.cleaned_data["date_from"]
        date_to = form.cleaned_data["date_to"]
        building = form.cleaned_data["building"]
        balance_type = form.cleaned_data["balance_type"]

        wb = build_transactions_workbook(
            date_from=date_from,
            date_to=date_to,
            building=building,
            balance_filter=balance_type,
        )

        suffix = _balance_suffix(balance_type)
        filename = (
            f"export_выдачи_{date_from:%Y-%m-%d}_{date_to:%Y-%m-%d}_{suffix}.xlsx"
        )

        return _make_excel_response(wb, filename)


def _apply_balance_filter(queryset, balance_type):
    if balance_type in ("balance", "balance_only"):
        return queryset.filter(on_balance=True)
    if balance_type in ("non_balance", "non_balance_only"):
        return queryset.filter(on_balance=False)
    return queryset


def _autosize_worksheet(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                value = str(cell.value or "")
            except Exception:
                value = ""
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)


def _style_header(ws, row_num=1):
    for cell in ws[row_num]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")


class ExportStocksExcelView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        form = ExportStocksForm(request.POST)
        if not form.is_valid():
            return HttpResponse("Некорректные параметры выгрузки.", status=400)

        balance_type = form.cleaned_data["balance_type"]

        wb = Workbook()
        ws = wb.active
        ws.title = "Сводный"

        summary_qs = (
            GlobalStock.objects
            .select_related("cartridge")
            .order_by("cartridge__vendor", "cartridge__code", "on_balance")
        )
        summary_qs = _apply_balance_filter(summary_qs, balance_type)

        ws.append([
            "Производитель",
            "Модель картриджа",
            "Описание",
            "Тип",
            "Количество",
        ])

        for row in summary_qs:
            ws.append([
                row.cartridge.vendor,
                row.cartridge.code,
                row.cartridge.title,
                "На балансе" if row.on_balance else "Не на балансе",
                row.qty,
            ])

        _style_header(ws)
        _autosize_worksheet(ws)

        buildings = Building.objects.order_by("name")
        for building in buildings:
            bws = wb.create_sheet(title=building.name[:31])

            bqs = (
                BuildingStock.objects
                .filter(building=building)
                .select_related("cartridge")
                .order_by("cartridge__vendor", "cartridge__code", "on_balance")
            )
            bqs = _apply_balance_filter(bqs, balance_type)

            bws.append([
                "Производитель",
                "Модель картриджа",
                "Описание",
                "Тип",
                "Количество",
            ])

            for row in bqs:
                bws.append([
                    row.cartridge.vendor,
                    row.cartridge.code,
                    row.cartridge.title,
                    "На балансе" if row.on_balance else "Не на балансе",
                    row.qty,
                ])

            _style_header(bws)
            _autosize_worksheet(bws)

        suffix = _balance_suffix(balance_type)
        filename = f"export_остатки_{suffix}.xlsx"

        return _make_excel_response(wb, filename)