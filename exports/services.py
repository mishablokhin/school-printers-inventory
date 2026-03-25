from collections import defaultdict
from datetime import datetime, time

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.utils import timezone

from inventory.models import StockTransaction, Building


THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def _safe_sheet_title(name: str) -> str:
    invalid = ['\\', '/', '*', '[', ']', ':', '?']
    for ch in invalid:
        name = name.replace(ch, " ")
    return name[:31] or "Лист"


def _apply_header_style(cell):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4472C4")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER


def _apply_body_style(cell):
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.border = THIN_BORDER


def _set_column_widths(ws):
    widths = {
        1: 18,  # Дата
        2: 14,  # Корпус-склад
        3: 14,  # Корпус назначения
        4: 12,  # Кабинет
        5: 30,  # Ответственный
        6: 24,  # Картридж
        7: 16,  # Баланс
        8: 10,  # Кол-во
        9: 34,  # Модель принтера
        10: 18, # Инв. номер
        11: 30, # Оформил
        12: 40, # Комментарий
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _build_queryset(date_from, date_to, building=None, balance_filter="all"):
    dt_from = timezone.make_aware(datetime.combine(date_from, time.min))
    dt_to = timezone.make_aware(datetime.combine(date_to, time.max))

    qs = (
        StockTransaction.objects
        .select_related("created_by", "cartridge", "building", "printer", "printer__room", "printer__printer_model")
        .filter(
            tx_type=StockTransaction.Type.OUT,
            created_at__gte=dt_from,
            created_at__lte=dt_to,
        )
        .order_by("created_at", "id")
    )

    if building:
        qs = qs.filter(building_snapshot=building.name)

    if balance_filter == "balance_only":
        qs = qs.filter(on_balance=True)
    elif balance_filter == "non_balance_only":
        qs = qs.filter(on_balance=False)

    return qs


def _group_transactions_by_building(qs, selected_building=None):
    grouped = defaultdict(list)

    if selected_building:
        grouped[selected_building.name] = list(qs)
        return grouped

    for tx in qs:
        sheet_name = tx.building_snapshot or (tx.building.name if tx.building else "Без корпуса")
        grouped[sheet_name].append(tx)

    return grouped


def _tx_author_name(tx):
    user = tx.created_by
    full_name = f"{user.last_name} {user.first_name}".strip()
    return full_name or user.username


def _tx_balance_label(tx):
    return "На балансе" if tx.on_balance else "Не на балансе"


def build_transactions_workbook(*, date_from, date_to, building=None, balance_filter="all"):
    qs = _build_queryset(
        date_from=date_from,
        date_to=date_to,
        building=building,
        balance_filter=balance_filter,
    )

    grouped = _group_transactions_by_building(qs, selected_building=building)

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    headers = [
        "Дата и время",
        "С какого склада",
        "Корпус назначения",
        "Кабинет",
        "Кому выдано",
        "Картридж",
        "Тип учёта",
        "Количество",
        "Модель принтера",
        "Инвентарный номер",
        "Оформил",
        "Комментарий",
    ]

    if not grouped:
        ws = wb.create_sheet("Выгрузка")
        ws.append(headers)
        for cell in ws[1]:
            _apply_header_style(cell)

        ws.append(["Нет данных за выбранный период"])
        _set_column_widths(ws)
        return wb

    for building_name, txs in grouped.items():
        ws = wb.create_sheet(_safe_sheet_title(building_name))

        ws.append(headers)
        for cell in ws[1]:
            _apply_header_style(cell)

        for tx in txs:
            ws.append([
                timezone.localtime(tx.created_at).strftime("%d.%m.%Y %H:%M"),
                tx.building.name if tx.building else "",
                tx.building_snapshot or "",
                tx.room_snapshot or "",
                tx.issued_to_snapshot or tx.issued_to or "",
                f"{tx.cartridge.vendor} {tx.cartridge.code}",
                _tx_balance_label(tx),
                tx.qty,
                tx.printer_model_snapshot or "",
                tx.printer_inventory_tag_snapshot or "",
                _tx_author_name(tx),
                tx.comment or "",
            ])

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                _apply_body_style(cell)

        ws.freeze_panes = "A2"
        _set_column_widths(ws)

    return wb